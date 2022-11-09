"""
資料來源:
https://selenium-python-zh.readthedocs.io/en/latest/getting-started.html

"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment

# 初始設定
wb = Workbook()
sheet = wb.active

option = webdriver.ChromeOptions()          # google chrom
driver = webdriver.Chrome('chromedriver.exe', chrome_options=option)
driver.get('https://www.twse.com.tw/zh/page/trading/exchange/MI_INDEX.html')

time.sleep(2)               # 休息2秒


# 找資料 (年,月,日,哪個廠業的代碼,工作表單數字)
def findData(yyyy, mmm, ddd, industryNum, sheetNum):
    # 日期設定
    year = driver.find_element(by=By.NAME, value="yy")
    year.send_keys(yyyy)
    month = driver.find_element(by=By.NAME, value="mm")
    month.send_keys(mmm)
    day = driver.find_element(by=By.NAME, value="dd")
    day.send_keys(ddd)

    # 資料
    data = Select(driver.find_element(by=By.NAME, value="type"))
    index = industryNum
    data.select_by_index(index)

    # 按下 搜尋鈕
    search = driver.find_element(by=By.CLASS_NAME, value="button")
    search.click()
    time.sleep(1)

    # 每頁顯示全部資料
    dataNum = driver.find_element(by=By.NAME, value="report-table1_length")
    dataNum.send_keys("全部")
    time.sleep(1)

    # 將所有的HTML的程式全部轉成BS4
    str1 = driver.page_source
    soup = BeautifulSoup(str1, "html.parser")

    AllData = {}    # 放入所有資料的地方
    conNum = 0      # 控制所有資料的變數

    # 找出 資料SHOW 出來的地方
    firstData = soup.select(".show")[0]

    # 找出 資料的標題
    AllData[conNum] = firstData.select("#subtitle1")[0].string
    conNum += 1

    # 顯示資料的表格
    tabelData = firstData.select("#report-table1")[0]

    # 一排一排找標頭
    def findTitle(className):
        ans = []
        for row in tabelData.select(className):
            ans.append(row.string)
        return ans

    # 找標頭1
    AllData[conNum] = findTitle(".group")
    conNum += 1
    # 找標頭2
    AllData[conNum] = findTitle(".sorting_disabled")
    conNum += 1

    data1 = tabelData.select(".odd")        # 資料所在的地方1
    data2 = tabelData.select(".even")       # 資料所在的地方2

    # 利用 資料所在的地方長度 找資料
    for x in range(0, len(data1)):
        dataFind1 = []
        dataFind2 = []
        # 找出　資料內的文字
        for y in range(0, len(data1[x].select(".dt-head-center"))):
            dataFind1.append(data1[x].select(".dt-head-center")[y].string)
            # 如果 兩邊數量相同 則沒有事情 繼續讀取資料
            if len(data1) != len(data2):
                # 當資料長度不同時,讓比較少的陣列 不進入 最後一筆
                if x == (len(data1)-1):
                    continue        # for 迴圈繼續
                else:
                    dataFind2.append(data2[x].select(".dt-head-center")[y].string)
            # 一般狀況繼續讀取資料
            else:
                dataFind2.append(data2[x].select(".dt-head-center")[y].string)

        AllData[conNum] = dataFind1
        conNum += 1
        # 如果 兩邊數量相同 則沒有事情 繼續將資料放入
        if len(data1) != len(data2):
            if x == (len(data1) - 1):
                break  # 離開迴圈
            else:
                AllData[conNum] = dataFind2
                conNum += 1
        # 相同情況 放入最後一筆資料
        else:
            AllData[conNum] = dataFind2
            conNum += 1

    def dataPrint(dataAll):
        for x in range(0, len(dataAll)):
            print(dataAll[x])

    def saveXlsx(dataIn, sheetNum):
        # 工作表單名稱
        name = yyyy + mmm + ddd

        # 新增日期名的工作表
        wb.create_sheet(title=name)

        # 資料輸入到指定的工作表
        wb.active = sheetNum
        sheet = wb.active

        # 資料樣式
        style = Alignment(horizontal="center", vertical="center")

        # 將儲存格合併
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=len(dataIn[2]))

        # 輸入第一個標頭資料
        sheet.cell(row=1, column=1).value = dataIn[0]

        # 將儲存格合併
        sheet.merge_cells(start_row=2, end_row=2, start_column=1, end_column=11)
        sheet.merge_cells(start_row=2, end_row=2, start_column=12, end_column=16)

        # 輸入第二個標頭資料
        sheet.cell(row=2, column=1).value = dataIn[1][0]
        sheet.cell(row=2, column=12).value = dataIn[1][1]

        # 檔案資料丟進excel
        for x in range(2, len(dataIn)):
            sheet.append(dataIn[x])

        # 將所有資料置中
        for x in range(1, sheet.max_row+1):
            for y in range(1, sheet.max_column+1):
                sheet.cell(row=x, column=y).alignment = style

    # 進入印資料
    dataPrint(AllData)

    # 進入存檔
    saveXlsx(AllData, sheetNum)


conSheet = 1    # 控制資料輸入到哪個工作表的變數
for Day in range(25, 29):
    findData("2022", "04", str(Day), 35, conSheet)
    conSheet += 1

"""
將 excel 預設的工作表 刪除
必須先給他隨意的資料 不然 系統會找不到
然後 將他改名 改完之後 刪除
沒改系統找不到
"""
sheet.cell(row=1, column=1).value = "test"
sheet.title = "test"
delSheet = wb["test"]
wb.remove(delSheet)

# 儲存資料
wb.save("Alldata.xlsx")
print("---------------------------\nxlsx檔 儲存成功!!")

# 關掉視窗
driver.close()
